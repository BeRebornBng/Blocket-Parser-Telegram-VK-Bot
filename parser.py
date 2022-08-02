import os.path
import threading
from googletrans import Translator
from time import sleep
import re
import platform
from selenium.webdriver.common.by import By
from selenium.webdriver.common.action_chains import ActionChains
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from openpyxl import load_workbook
from telebot.types import InputMediaPhoto
from psycopg2 import Error
from telebot import types
import requests
import schedule
import xmltodict
import datetime
import configparser
import telebot
import time
from collections import deque
from VK import AddImageToDirectory, SendPostToVk, AddImageToVkAlbum

config = configparser.ConfigParser()
config.read('Config.ini')
token = str(config['Telegram']['token'])
linux_flag = False

bot = telebot.TeleBot(token)

data_list = deque()
data_count = -1
data_flag = False
massLinks = []
getYachtSleep = int(config['Program']['getYachtSleep'])
btnSleep = int(config['Program']['btnSleep'])


@bot.message_handler(commands=['start'])
def start_message(message):
    bot.send_message(message.chat.id, "Привет ✌️ ")


bot.polling()


def translate(text=None):
    translator = Translator()
    translation = translator.translate(text, dest='ru')
    return translation.text


def move_to_element(element, driver):
    ActionChains(driver) \
        .scroll_to_element(element) \
        .perform()


def getYachtName(driver):
    global getYachtSleep
    try:
        #'//*[@id="skip-tabbar"]/div[2]/div[2]/div[1]/div/article/div[1]/div[2]/h1'
        if(linux_flag):
            yachtNameElement = driver.find_element(By.CSS_SELECTOR, '#skip-tabbar > div:nth-child(2) > article > div.LoadingAnimationStyles__PlaceholderWrapper-sc-c75se8-0.bXGjLh > div.Hero__ContentWrapper-sc-1mjgwl-3.hWIxa-D > h1')
        else:
            yachtNameElement = driver.find_element(By.CSS_SELECTOR,
                                               '#skip-tabbar > div.MediumLayout__CenterWithPadding-sc-y8zw9h-4.MediumLayout__CenterWithPaddingAndBg-sc-y8zw9h-5.hwvqhT.cdvLMa > div.MediumLayout__BodyWrapper-sc-y8zw9h-0.knswkO > div.MediumLayout__BodyLeft-sc-y8zw9h-2.kTyLZD > div > article > div.LoadingAnimationStyles__PlaceholderWrapper-sc-c75se8-0.bXGjLh > div.Hero__ContentWrapper-sc-1mjgwl-3.hWIxa-D > h1')
        # смещение до элемента yachtNameElement
        move_to_element(yachtNameElement, driver)
        sleep(getYachtSleep)
        return translate(str(yachtNameElement.text))
    except (Exception, Error) as error:
        print('Название яхты не найдено')
        # название яхты не найдено
        return ''


def getYachtPrice(driver):
    global getYachtSleep
    try:
        #'//*[@id="skip-tabbar"]/div[2]/div[2]/div[1]/div/article/div[1]/div[2]/div[2]'
        if(linux_flag):
            yachtPrice = driver.find_element(By.CSS_SELECTOR, '#skip-tabbar > div:nth-child(2) > article > div.LoadingAnimationStyles__PlaceholderWrapper-sc-c75se8-0.bXGjLh > div.Hero__ContentWrapper-sc-1mjgwl-3.hWIxa-D > div.Hero__PriceWrapper-sc-1mjgwl-5.hGqpgd > div')
        else:
            yachtPrice = driver.find_element(By.CSS_SELECTOR,
                                         '#skip-tabbar > div.MediumLayout__CenterWithPadding-sc-y8zw9h-4.MediumLayout__CenterWithPaddingAndBg-sc-y8zw9h-5.hwvqhT.cdvLMa > div.MediumLayout__BodyWrapper-sc-y8zw9h-0.knswkO > div.MediumLayout__BodyLeft-sc-y8zw9h-2.kTyLZD > div > article > div.LoadingAnimationStyles__PlaceholderWrapper-sc-c75se8-0.bXGjLh > div.Hero__ContentWrapper-sc-1mjgwl-3.hWIxa-D > div.Hero__PriceWrapper-sc-1mjgwl-5.hGqpgd > div.TextHeadline1__TextHeadline1Wrapper-sc-1bi3cli-0.bIxKdL.Price__StyledPrice-sc-crp2x0-0.kIhjJa')
        # смещение до элемента yachtPrice
        move_to_element(yachtPrice, driver)
        sleep(getYachtSleep)
        return str(yachtPrice.text)
    except (Exception, Error) as error:
        print("Цена яхты не найдена")
        # цена яхты не найдена
        return ''


def getYachDescription(driver):
    global getYachtSleep
    try:
        #'//*[@id="skip-tabbar"]/div[2]/div[2]/div[1]/div/article/div[2]/div[2]/div[3]/div[2]'
        if(linux_flag):
            yachtDescription = driver.find_element(By.CSS_SELECTOR, '#skip-tabbar > div:nth-child(2) > article > div.AdMotor__Details-sc-934l4e-1.fnuIsR > div:nth-child(2) > div.BodyCard__StyledCard-sc-15r463q-0.dSLbnj > div.ExpandableContent__Content-sc-11a0rym-0.eUvFOR > div')
        else:
            yachtDescription = driver.find_element(By.CSS_SELECTOR,
                                               '#skip-tabbar > div.MediumLayout__CenterWithPadding-sc-y8zw9h-4.MediumLayout__CenterWithPaddingAndBg-sc-y8zw9h-5.hwvqhT.cdvLMa > div.MediumLayout__BodyWrapper-sc-y8zw9h-0.knswkO > div.MediumLayout__BodyLeft-sc-y8zw9h-2.kTyLZD > div > article > div.AdMotor__Details-sc-934l4e-1.fnuIsR > div:nth-child(2) > div.BodyCard__StyledCard-sc-15r463q-0.dSLbnj > div.ExpandableContent__Content-sc-11a0rym-0.eUvFOR > div')
        # смещение до элемента yachtDescription
        move_to_element(yachtDescription, driver)
        sleep(getYachtSleep)
        return translate(yachtDescription.text)
    except (Exception, Error) as error:
        # описание яхты не найдено
        print("Описание яхты не найдено")
        return ''


def getCountImages(driver):
    numbers = driver.find_element(by=By.XPATH,
                                  value='//*[@id="skip-tabbar"]/div[2]/div[2]/div[1]/div/article/div[1]/div[1]/div/div[2]/span').text
    numbers = re.findall('\d+', numbers)
    return int(numbers[1])


def getYachtImages(driver):
    global getYachtSleep
    global btnSleep
    try:
        yachtNameElement = driver.find_element(By.XPATH,
                                               '//*[@id="skip-tabbar"]/div[2]/div[2]/div[1]/div/article/div[1]/div[2]/h1')
        # смещение до элемента yachtNameElement
        move_to_element(yachtNameElement, driver)
        sleep(getYachtSleep)
        start = 1
        end = getCountImages(driver)
        media = []
        t = None
        while start < end + 1:
            try:
                t = driver.find_element(by=By.XPATH,
                                        value=f'//*[@id="skip-tabbar"]/div[2]/div[2]/div[1]/div/article/div[1]/div[1]/div/div[1]/div/div[{start}]').get_attribute(
                    'style')
            except (Exception, Error) as error:
                print('Не получен атрибут style')
                print(error)
            try:
                driver.find_element(by=By.XPATH,
                                    value='//*[@id="skip-tabbar"]/div[2]/div[2]/div[1]/div/article/div[1]/div[1]/div/button[1]').click()
            except (Exception, Error) as error:
                print("Кнопка перехода на другую картинку не нажата")
            sleep(btnSleep)
            tp = t.split('"')
            tp2 = tp[1].split('"')
            tp2 = tp2[0].replace('webp', 'png')
            # media.append(InputMediaPhoto(tp2))
            media.append(tp2)
            start += 1
        return media
    except (Exception, Error) as error:
        # фото яхты не найдено
        print("ФОто яхты не найдено")
        return


def Page(driver, page):
    if (page <= 1):
        return 1
    if (driver.current_url == 'https://www.blocket.se/annonser/hela_sverige/fordon/batar/segelbat?cg=1062&page=1'):
        return 1
    else:
        driver.get(f'https://www.blocket.se/annonser/hela_sverige/fordon/batar/segelbat?cg=1062&page={page}')
        return 0


def findLink(sh, driver):
    try:
        for s in sh.values:
            t = str(s)
            tp = t.split("'")
            tp2 = tp[1].split("'")
            if (driver.current_url == tp2[0]):
                print("Здесь уже был")
                return 1
    except (Exception, Error) as error:
        print("Пусто")


def convertRubtoCron(driver):
    try:
        response = requests.get(
            'https://www.cbr.ru/scripts/XML_daily.asp?date_req=' + datetime.datetime.now().strftime('%d/%m/%Y'))
        priceCron = getYachtPrice(driver)
        temp = ''
        for price in priceCron:
            if (price == ' '):
                continue
            else:
                temp += price
        temp = re.findall('\d+', temp)
        priceRub = str(xmltodict.parse(response.text)['ValCurs']['Valute'][29]['Value']).replace(',', '.')
        return (float(temp[0]) / 10 * float(priceRub))
    except (Exception, Error) as error:
        print(error)
        return ''


def CountPages(driver):
    count = int(config['Program']['page'])
    while True:
        driver.get(f'https://www.blocket.se/annonser/hela_sverige/fordon/batar/segelbat?cg=1062&page={count}')
        sleep(3)
        if (
                str(driver.current_url) != f'https://www.blocket.se/annonser/hela_sverige/fordon/batar/segelbat?cg=1062&page={count}'):
            driver.get(f'https://www.blocket.se/annonser/hela_sverige/fordon/batar/segelbat?cg=1062&page={count - 2}')
            print('return')
            return count - 2
        count += 1


def getLinksFormatToTelegram(links):
    media = []
    for link in links:
        media.append(InputMediaPhoto(link))
    return media


def getDataDict(driver):
    global data_count
    data = {
        'link': '',
        'name': '',
        'price': '',
        'description': '',
        'imagelinks': '',
        'vkLinks': ''
        'userLink'
    }
    links = getYachtImages(driver)
    data['link'] = str(driver.current_url)
    massLinks.append(driver.current_url)
    data['userLink'] = str(config['User']['UserLink']) + '?utm_content=' + str(driver.current_url).split('https://www.blocket.se/annons/')[1]
    print(data['link'])
    data['imagelinks'] = getLinksFormatToTelegram(links)
    data['name'] = getYachtName(driver)
    price = ("{:.2f}".format((convertRubtoCron(driver))))
    data['price'] = getYachtPrice(driver) + '\n' + price + ' руб'
    data['description'] = getYachDescription(driver)
    data['vkLinks'] = links
    data_count += 1
    return data


def findLinkInDict(driver):
    for link in massLinks:
        if link == str(driver.current_url):
            return 1
    return 0


def Parser2():
    global linux_flag
    global data_count
    global data_flag
    global btnSleep
    driver = None
    if(platform.system() == 'Windows'):
        service = Service('chromedriver.exe')
        options = webdriver.ChromeOptions()
        options.headless = False
        options.add_argument('log-level=3')
        driver = webdriver.Chrome(
            service=service, options=options)
    else:
        try:
            linux_flag = True
            service = Service(executable_path=os.path.abspath('chromedriver'))
            options = webdriver.ChromeOptions()
            options.headless = True
            driver = webdriver.Chrome(service=service, options=options)
        except (Exception, Error) as error:
            print("ошибка")
            print(error)
    driver.get('https://www.blocket.se/annonser/hela_sverige/fordon/batar/segelbat?cg=1062&page=1')
    sleep(3)
    driver.find_element(by=By.XPATH, value='//*[@id="accept-ufti"]').click()
    sleep(btnSleep)
    page = None
    if (config['Program']['flag'])=='True':
        page = CountPages(driver)
    elif config['Program']['flag']=='False':
        page = int(config['Program']['page'])
        driver.get(f'https://www.blocket.se/annonser/hela_sverige/fordon/batar/segelbat?cg=1062&page={page}')
    pageIndex = int(config['Program']['PageIndex'])
    while True:
        sleep(2)
        driver.execute_script("window.scrollBy(0,7000)", "")
        sleep(4)
        while (True):
            print("Страница = " + str(page))
            try:
                if (pageIndex == -1):
                    print('Достигнут конец страницы')
                    break
                else:
                    element = driver.find_element(by=By.CSS_SELECTOR,
                                                  value=f'#__next > div > main > div.MediumLayout__BodyWrapper-sc-q6qal1-2.bCbTFf > div.MediumLayout__BodyLeft-sc-q6qal1-3.cOXCVa > div:nth-child(3) > div > div:nth-child({pageIndex}) > article > div.styled__Content-sc-1kpvi4z-2.hqmvlX > div.styled__SubjectWrapper-sc-1kpvi4z-15.bJucZO > h2 > a')
                    sleep(1)
                    move_to_element(element, driver)
                    sleep(2)
                    element.click()
                    print('Ссылка по счёту = ' + str(pageIndex))
                    sleep(3)
                    wb = load_workbook('output.xlsx')
                    sh1 = wb['List']
                    sh = wb.active
                    print("Помещено объектов = " + str(len(data_list)))
                    if (findLink(sh, driver)):
                        print('Эта ссылка уже записана')
                        driver.back()
                        sleep(1.5)
                        pageIndex -= 1
                        continue
                    try:
                        if (findLinkInDict(driver)):
                            print("Совпало")
                            driver.back()
                            sleep(1.5)
                            pageIndex -= 1
                            continue

                        data_list.append(getDataDict(driver))
                        sleep(0.5)
                        driver.back()
                        sleep(2)
                        pageIndex -= 1
                        continue
                    except (Exception, Error) as error:
                        print("ERRORXXX")
                        print(error)
                        sleep(1)
                        driver.back()
                        sleep(1)
                        pageIndex -= 1
                        continue
            except (Exception, Error) as error:
                print("Не нашло ссылку на яхту")
                pageIndex -= 1
                continue
                # Ошибка, не нашло ссылку на яхту.
        page -= 1
        if (Page(driver, page)):
            data_flag = True
            pageIndex = int(config['Program']['PageIndex'])
            page = 1
            driver.get('https://www.blocket.se/annonser/hela_sverige/fordon/batar/segelbat?cg=1062&page=1')
            sleep(3)
            continue
        pageIndex = int(config['Program']['PageIndex'])


def SendTelegram():
    wb = load_workbook('output.xlsx')
    sh1 = wb['List']
    sh = wb.active
    global data_count
    global data_flag
    print(data_flag)
    print("SENDMESSAGE DATA LIST")
    if(len(data_list) >1):
        if data_flag:
            data = data_list.pop()
            if(data['name']=='' and data['description']=='' and len(data['imagelinks'] or data['price']=='')==0):
                print("Empty")
            else:
                try:
                    data = data_list.pop()
                    msg = str(str(data['userLink']) + '\n' + data['name'] + '\n' + data['price'] + '\n' + data['description'])
                    try:
                        start = 0
                        for link in data['vkLinks']:
                            if(start==6):
                                break
                            AddImageToDirectory(start, link)
                            start += 1
                        SendPostToVk('Заказать перегон яхты '+ msg)
                        print("Пост в вк успешно отправлен")
                    except (Exception, Error) as error:
                        print(error)
                        print("Не отправился пост в вк")
                    keyboard = types.InlineKeyboardMarkup()
                    url_btn = types.InlineKeyboardButton(text=('ЗАКАЗАТЬ ПЕРЕГОН ЯХТЫ'), url=data['userLink'])
                    keyboard.add(url_btn)
                    bot.send_media_group(config['Telegram']['channel'], data['imagelinks'])
                    time.sleep(10)
                    bot.send_message(config['Telegram']['channel'], msg, reply_markup=keyboard)
                    data_count -= 1
                    print('Пост в телеграм успешно отправлен')
                    try:
                        t = data['link']
                        sh.append([t])
                        wb.save('output.xlsx')
                        print('Записал ссылку')
                    except (Exception, Error) as error:
                        print(error)
                        print("Не записал ссылку")
                except (Exception, Error) as error:
                    print(error)
                    print("Ошибка в блоке SendMessage")


def thr():
    while True:
        schedule.run_pending()
        time.sleep(1)


if __name__ == "__main__":
    try:
        print('Start a program')
        threading.Thread(target=Parser2).start()
        if(str(config['Program']['sendMess'])=='True'):
            schedule.every(30).seconds.do(SendTelegram)
        schedule.every().days.at((config['TimeUP']['time1'])).do(SendTelegram)
        schedule.every().days.at((config['TimeUP']['time2'])).do(SendTelegram)
        schedule.every().days.at((config['TimeUP']['time3'])).do(SendTelegram)
        schedule.every().days.at((config['TimeUP']['time4'])).do(SendTelegram)
        schedule.every().days.at((config['TimeUP']['time5'])).do(SendTelegram)
        threading.Thread(target=thr).start()
    except (Exception, Error) as error:
        print("MAIN ERROR")
        print(error)